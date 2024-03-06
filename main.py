import os
import json
import pandas as pd
from openpyxl import load_workbook
from PyPDF2 import PdfReader
from docx import Document

def excel_to_dataframe(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)
    return pd.DataFrame(data)

def pdf_to_dataframe(file_path):
    with open(file_path, 'rb') as file:
        reader = PdfReader(file)
        data = []
        for page in reader.pages:
            data.append(page.extract_text())
    return pd.DataFrame(data, columns=["Text"])

def word_to_dataframe(file_path):
    doc = Document(file_path)
    data = []
    for paragraph in doc.paragraphs:
        data.append(paragraph.text)
    return pd.DataFrame(data, columns=["Text"])

def file_to_dataframe(file_path):
    _, file_extension = os.path.splitext(file_path)
    if file_extension.lower() == '.xlsx':
        return excel_to_dataframe(file_path)
    elif file_extension.lower() == '.pdf':
        return pdf_to_dataframe(file_path)
    elif file_extension.lower() == '.docx':
        return word_to_dataframe(file_path)
    else:
        raise ValueError("Unsupported file format")

def main():
    input_file_path = input("Lütfen dönüştürmek istediğiniz dosyanın yolunu girin: ")

    df = file_to_dataframe(input_file_path)

    output_file_name = os.path.splitext(os.path.basename(input_file_path))[0] + ".json"
    output_file_path = os.path.join(os.path.dirname(input_file_path), output_file_name)

    df.to_json(output_file_path, orient="records")

    with open(output_file_path, 'r', encoding='utf-8') as json_file:
        json_data = json.load(json_file)

    with open(output_file_path, 'w', encoding='utf-8') as json_file:
        json.dump(json_data, json_file, ensure_ascii=False, indent=4)

    print(f"Dönüştürülen veriler '{output_file_name}' dosyasına başarıyla kaydedildi.")

if __name__ == "__main__":
    main()
